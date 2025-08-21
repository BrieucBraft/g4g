import cv2
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import dropbox
import requests
from tkinter import messagebox
import os
import re
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
import csv
import json
from app import config

def convert_xlsx_to_csv(input_file: str, output_file: str):
    try:
        df = pd.read_excel(input_file)
        if 'test passed' in df.columns:
            df['test passed'] = ''
        df.to_csv(output_file, sep=';', index=False, encoding='iso8859_15')
        print(f"File successfully converted and saved to {output_file}")
    except Exception as e:
        print(f"An error occurred: {e}")

def sort_and_replace(input_excel, output_csv):
    df_excel = pd.read_excel(input_excel)
    df_csv = pd.read_csv(output_csv, sep=';', dtype=str, encoding='iso8859_15')
    df_excel_sorted = df_excel.sort_values(by='unique ID').reset_index(drop=True)
    df_csv_sorted = df_csv.sort_values(by='unique ID').reset_index(drop=True)
    df_csv_sorted['index m'] = df_excel_sorted['index m'].values
    df_csv_sorted.to_csv(output_csv, sep=';', index=False, encoding='iso8859_15')

def move_columns_right(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    headers = {cell.value: cell.column for cell in sheet[1]}
    index_m_col = headers['index m']
    index_m1_col = headers['index m-1']
    index_m2_col = headers['index m-2']
    test_passed_col = headers['test passed']
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        test_passed_value = row[test_passed_col - 1].value
        if test_passed_value is None or test_passed_value == '':
            row[index_m2_col - 1].value = row[index_m1_col - 1].value
            row[index_m1_col - 1].value = row[index_m_col - 1].value
            row[index_m_col - 1].value = ''
    workbook.save(file_path)
    print(f"Processing complete. Modified file saved as {file_path}")

def replace_comma_to_dot_separator(input_file):
    df = pd.read_csv(input_file, sep=';', dtype=str, encoding='iso8859_15')
    df['index m-1'] = df['index m-1'].str.replace(',', '.', regex=False)
    df['index m-2'] = df['index m-2'].str.replace(',', '.', regex=False)
    df['index m'] = df['index m'].str.replace(',', '.', regex=False)
    df['index extrapole'] = df['index extrapole'].str.replace(',', '.', regex=False)
    df['index I+'] = df['index I+'].str.replace(',', '.', regex=False)
    df['test passed'] = df['test passed'].str.replace('HYPERLINK', 'LIEN_HYPERTEXTE', regex=False)
    df.to_csv(input_file, sep=';', index=False, encoding='iso8859_15')

def replace_dot_to_comma_separator(input_file):
    df = pd.read_csv(input_file, sep=';', dtype=str, encoding='iso8859_15')
    df['index m-1'] = df['index m-1'].str.replace('.', ',', regex=False)
    df['index m-2'] = df['index m-2'].str.replace('.', ',', regex=False)
    df['index m'] = df['index m'].str.replace('.', ',', regex=False)
    df['index extrapole'] = df['index extrapole'].str.replace('.', ',', regex=False)
    df['index I+'] = df['index I+'].str.replace('.', ',', regex=False)
    df['test passed'] = df['test passed'].str.replace('LIEN_HYPERTEXTE', 'HYPERLINK', regex=False)
    df['test passed'] = df['test passed'].str.replace(';', ',', regex=False)
    df.to_csv(input_file, sep=';', index=False, encoding='iso8859_15')

def combine_images(image_dict):
    for key, image_paths in image_dict.items():
        if len(image_paths) == 1:
            img = Image.open(image_paths[0])
            img.save(image_paths[0])
        elif len(image_paths) == 2:
            img1 = Image.open(image_paths[0])
            img2 = Image.open(image_paths[1])
            if img1.height != img2.height:
                new_height = min(img1.height, img2.height)
                img1 = img1.resize((int(img1.width * new_height / img1.height), new_height))
                img2 = img2.resize((int(img2.width * new_height / img2.height), new_height))
            total_width = img1.width + img2.width
            combined_img = Image.new('RGB', (total_width, img1.height))
            combined_img.paste(img1, (0, 0))
            combined_img.paste(img2, (img1.width, 0))
            combined_img.save(image_paths[0])
            combined_img.save(image_paths[1])

def isInIds(identite, existing_ids):
    if identite in existing_ids:
        return True, identite
    count = 0
    for exist_id in existing_ids:
        if identite in exist_id:
            count += 1
    if count == 1:
        return True, exist_id
    return False, identite

def checkID(identite, data):
    existing_ids = [row['unique ID'] for row in data]
    count = 0
    for id in existing_ids:
        if identite in id:
            count +=1
            remember = id
    if count == 1:
        identite = remember
    return identite

def add_spaces_before_capitals(input_string):
    input_string = input_string.replace(" ", "")
    spaced_string = re.sub(r'(?<!^)(?<![A-Z])([A-Z]+)', r' \1', input_string)
    return spaced_string.strip()

def get_info_from_name(image_file):
    identite = None
    code = None
    nom_compteur = None
    date = None
    N = None
    PC = None
    newName = image_file.split('.')[0]
    code = newName[:7]
    newName = newName[7:]
    newName = add_spaces_before_capitals(newName)
    numTir = newName.count("-")
    if numTir == 2:
        index = newName.find('-')
        N = newName[index+1]
        newName = newName.replace(f'-{N}', '', 1)
        index = newName.find('-')
        date = newName[index+1:]
        newName = newName[:index]
    elif numTir == 1:
        index = newName.find('-')
        date = newName[index+1:]
        newName = newName[:index]
    if '=' in newName:
        index = newName.find('=')
        PC = newName[index+1]
        newName = newName.replace(f'={PC}', '')
    nom_compteur = newName.strip()
    identite = code + ' ' + nom_compteur
    if N is not None:
        nom_compteur = nom_compteur + '-' + N
        identite = identite + '-' + N
    if PC is not None:
        identite = identite + '=' + PC
    return identite, code, nom_compteur, date

def open_csv_file(file_path):
    file_path = os.path.abspath(file_path)
    if os.path.exists(file_path):
        try:
            os.startfile(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Unable to open file: {e}")
    else:
        messagebox.showerror("File Not Found", f"The file '{file_path}' does not exist.")

def round_to_nearest_quarter_hour(dt):
    minutes = dt.minute
    hour = dt.hour
    if minutes % 15 != 0:
        quarter_hour = round(minutes / 15) * 15
        if quarter_hour == 60:
            quarter_hour = 0
            hour = hour + 1
        dt = dt.replace(hour = hour, minute=quarter_hour)
    return dt.replace(second=0, microsecond=0)

def mean_datetime(dates):
    datetimes = [datetime.strptime(date_str, "%Y-%m-%d %H:%M") for date_str in dates]
    total_seconds = sum((dt - datetimes[0]).total_seconds() for dt in datetimes)
    mean_seconds = total_seconds / len(datetimes)
    mean_dt = datetimes[0] + timedelta(seconds=mean_seconds)
    rounded_mean_dt = round_to_nearest_quarter_hour(mean_dt)
    return rounded_mean_dt.strftime("%Y-%m-%d %H:%M")

def translate_datetime(date_str):
    input_format = "%d_%m_%Y_%H_%M"
    output_format = "%Y-%m-%d %H:%M"
    dt = datetime.strptime(date_str, input_format)
    return dt.strftime(output_format)

def generate_available_dates(months_back=12, months_forward=3):
    today = datetime.today()
    available_dates = []
    for i in range(months_back):
        date = today - timedelta(days=i*30)
        formatted_date = date.strftime("%Y-%m")
        available_dates.append(formatted_date)
    for i in range(1, months_forward + 1):
        date = today + timedelta(days=i*30)
        formatted_date = date.strftime("%Y-%m")
        available_dates.append(formatted_date)
    available_dates = sorted(list(set(available_dates)))
    return available_dates

def get_access_token(app_key, app_secret, refresh_token):
    url = 'https://api.dropboxapi.com/oauth2/token'
    data = {'grant_type': 'refresh_token', 'refresh_token': refresh_token}
    auth = (app_key, app_secret)
    response = requests.post(url, data=data, auth=auth)
    response_data = response.json()
    if response.status_code == 200:
        return response_data['access_token']
    else:
        raise Exception(f"Error obtaining access token: {response_data}")

def is_earlier_or_empty(date_str, current_date):
    if date_str.strip() == '':
        return True
    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M")
    except:
        try:
            date_obj = datetime.strptime(date_str, "%d-%m-%y %H:%M")
        except:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    ym_obj = datetime.strptime(current_date, "%d_%m_%Y_%H_%M")
    return date_obj.date() < ym_obj.date()

def recursiveMoveValue(row, current_col, n=0, nmax = 2):
    if n == nmax or f'index m-{n+1}' not in row:
        return row
    next_col = f'index m-{n+1}'
    if current_col in row and row[current_col] != '':
        row = recursiveMoveValue(row, next_col, n + 1, nmax=2)
        if next_col in row:
            row[next_col] = row[current_col]
    row[current_col] = ''
    return row

def extrapolate_index(row, operatingHours, code, max_n=2):
    if code in operatingHours and 'Cogen' in row['NOM COMPTEUR']:
        try:
            dhm = float(operatingHours[code][0])-float(operatingHours[code][1])
            dhm1 = float(operatingHours[code][1])-float(operatingHours[code][2])
            dim1 = float(row['index m-1'])-float(row['index m-2'])
            im1 = float(row['index m-1'])
            if dhm <= 0 or dhm1 <= 0 or dim1 <= 0 or float(operatingHours[code][1]) == 0 or float(operatingHours[code][2]) == 0 or float(row['index m-2']) == 0:
                return None
            return str(round(im1+dhm*dim1/dhm1, 5))
        except:
            try:
                dim1 = float(row['index m-1'])-float(row['index m-2'])
                im1 = float(row['index m-1'])
                if dim1 <= 0 or float(row['index m-2']) == 0:
                    return None
                return str(round(im1 + dim1,5))
            except:
                return None
    else:
        x = []
        y = []
        for i in range(1, max_n+1):
            col_name = f'index m-{i}'
            if row.get(col_name) and row[col_name] != '':
                try:
                    y.append(float(row[col_name]))
                    x.append(-i)
                except ValueError:
                    continue
        if len(x) >= 2:
            if y[0] != 0 and y[1] != 0 and y[0] >= y[1]:
                slope, intercept = np.polyfit(x, y, 1)
                extrapolated_value = intercept
                return str(round(extrapolated_value,5))
            else:
                return None
        else:
            return None

def setDotPosition(value, row, column):
    index_I_plus = str(row.get(column, ''))
    if '.' in index_I_plus and index_I_plus != '0.0':
        dot_position = len(index_I_plus.split('.')[0])
        if '.' not in value:
            int_part = value[:dot_position]
            dec_part = value[dot_position:]
            newValue = f"{int_part}.{dec_part}"
        else:
            newValue = value
    elif len(index_I_plus) <= len(value):
        dot_position = len(index_I_plus)
        if '.' not in value:
            int_part = value[:dot_position]
            dec_part = value[dot_position:]
            newValue = f"{int_part}.{dec_part}"
        else:
            newValue = value
    else:
        newValue = value
    return newValue

def closestValue(values, extrapolatedStr, previous, row):
    extrapolated = row[extrapolatedStr]
    type = row['unique ID']
    try:
        values = [float(val) for val in values]
        closest = str(max(values))
        for value in values:
            try:
                if np.abs(float(value) - float(extrapolated)) < np.abs(float(closest)-float(extrapolated)) and float(value) >= float(previous):
                    closest = value
            except:
                continue
    except:
        return '0'
    return closest

def removeUnwanted(array, accepted):
    newA = []
    final = ""
    for elem in array:
        addTo = True
        for char in elem:
            if char not in accepted:
                addTo = False
        if addTo == True:
            newA.append(elem)
    return newA

def removePrefix(result):
    prefixes = ['180', '280', '200', '290', '100', '800']
    if len(result) >= 3:
        for i in range(len(result)-2):
            if result[i:i+3] in prefixes:
                return result.split(result[i:i+3])[1][:8]
        for prefix in prefixes:
            if prefix in result:
                return result.split(prefix)[1][:8]
        if result[:3] in prefixes:
            return result[3:]
    return result

def getROI(contours, img, rois, name, edges):
    roi = None
    max_area = np.shape(img)[0]*np.shape(img)[1]/150
    limit_area = np.shape(img)[0]*np.shape(img)[1]/15
    pixelMean = 120
    img_with_rectangles = img.copy()
    for cnt in contours:
        rect = cv2.minAreaRect(cnt)
        box = cv2.boxPoints(rect)
        box = np.intp(box)
        width = rect[1][0]
        height = rect[1][1]
        aspect_ratio = width / float(height) if height != 0 else 0
        area = width * height
        tempRoi = img[int(rect[0][1] - height/2):int(rect[0][1] + height/2),
                    int(rect[0][0] - width/2):int(rect[0][0] + width/2)]
        if 3 < aspect_ratio < 14 and max_area < area < limit_area and np.mean(tempRoi)<pixelMean:
            cv2.drawContours(img_with_rectangles, [box], 0, (0, 255, 0), 5)
            roi = tempRoi
            max_area = area
            boxCoord = {'x_min' : min(box[0][0], box[1][0])/1.1, 'y_min' : min(box[1][1], box[2][1])/1.1, 'x_max' : max(box[2][0], box[3][0])*1.1, 'y_max' : max(box[0][1], box[3][1])*1.1}
            print(aspect_ratio)
    rois.append(roi)
    try :
        plt.close('all')
        fig, axs = plt.subplots(1, 3, figsize=(12, 6))
        Verti = cv2.hconcat(rois)
        rect = cv2.cvtColor(img_with_rectangles, cv2.COLOR_BGR2RGB)
        axs[0].imshow(Verti, cmap='gray')
        axs[0].set_title('ROI')
        axs[0].axis('off')
        axs[1].imshow(rect)
        axs[1].set_title('Bounding rectangles')
        axs[1].axis('off')
        axs[2].imshow(edges, cmap='gray')
        axs[2].set_title('Edges Detected')
        axs[2].axis('off')
        plt.show()
        plt.figure()
        plt.imshow(cv2.hconcat(rois), cmap='gray')
        plt.axis('off')
        debug_path = os.path.join(config.CROPPED_IMAGES_DIR, name)
        plt.savefig(debug_path)
    except:
        pass
    return boxCoord

def toBox(x,y,w,h):
    p1 = [x,y+h]
    p2 = [x,y]
    p3 = [x+w, y]
    p4 = [x+w, y+h]
    return [p1,p2,p3,p4]

def list_dropbox_images(dropbox_folder_path):
    image_files = []
    DROPBOX_ACCESS_TOKEN = get_access_token(config.DROPBOX_APP_KEY, config.DROPBOX_APP_SECRET, config.DROPBOX_REFRESH_TOKEN)
    namespace_header = {".tag": "namespace_id", "namespace_id": config.DROPBOX_NAMESPACE_ID}
    headers = {"Dropbox-API-Path-Root": json.dumps(namespace_header)}
    dbx = dropbox.Dropbox(DROPBOX_ACCESS_TOKEN, headers=headers)
    try:
        result = dbx.files_list_folder(dropbox_folder_path)
        for entry in result.entries:
            if isinstance(entry, dropbox.files.FileMetadata):
                if entry.name.lower().endswith(('.jpg', '.jpeg', '.png')):
                    image_files.append(entry.name)
    except dropbox.exceptions.ApiError as err:
        print(f"Failed to list Dropbox folder: {err}")
    return image_files, dbx

def list_dropbox_folders(dropbox_folder_path, yearMonth):
    folder_list = []
    year = yearMonth.split('-')[0]
    month = yearMonth.split('-')[1]
    folder_pattern = re.compile(f'{dropbox_folder_path}' + r'/[A-Z]{3}-[A-Z0-9]{3}/[Pp][Hh][Oo][Tt][Oo][Ss]?/' + f'{year}-{month}' + r'$')
    DROPBOX_ACCESS_TOKEN = get_access_token(config.DROPBOX_APP_KEY, config.DROPBOX_APP_SECRET, config.DROPBOX_REFRESH_TOKEN)
    namespace_header = {".tag": "namespace_id", "namespace_id": config.DROPBOX_NAMESPACE_ID}
    headers = {"Dropbox-API-Path-Root": json.dumps(namespace_header)}
    dbx = dropbox.Dropbox(DROPBOX_ACCESS_TOKEN, headers=headers)
    try:
        result = dbx.files_list_folder(dropbox_folder_path, recursive=True)
        while True:
            print(result)
            for entry in result.entries:
                if isinstance(entry, dropbox.files.FolderMetadata):
                    if folder_pattern.match(entry.path_display):
                        folder_list.append(entry.path_display)
            if result.has_more:
                result = dbx.files_list_folder_continue(result.cursor)
            else:
                break
    except dropbox.exceptions.ApiError as err:
        print(f"Failed to list Dropbox folder: {err}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
    return folder_list

def download_image_from_dropbox(dropbox_path, local_path, dbx):
    try:
        dbx.files_download_to_file(local_path, dropbox_path)
    except dropbox.exceptions.ApiError as err:
        print(f"Failed to download {dropbox_path}: {err}")